import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import SignUpForm, AddRecordForm
from .models import Record

def home(request):
	records = Record.objects.all()
	# Check to see if logging in
	if request.method == 'POST':
		username = request.POST['username']
		password = request.POST['password']
		# Authenticate
		user = authenticate(request, username=username, password=password)
		if user is not None:
			login(request, user)
			messages.success(request, "Ha iniciado sesión!")
			return redirect('home')
		else:
			messages.success(request, "Se ha producido un error al iniciar sesión, inténtelo de nuevo....")
			return redirect('home')
	else:
		return render(request, 'home.html', {'records':records})


def logout_user(request):
    logout(request)
    messages.success(request, "Se cerro la sesión")
    return redirect('home')


def register_user(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            # Authenticate and Login
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            user = authenticate(username=username, password=password)
            login(request, user)
            messages.success(request, "Se ha registrado correctamente")
            return redirect('home')
    else:
        form = SignUpForm()
        return render(request, 'register.html', {'form':form})
    return render(request, 'register.html', {'form':form})

def customer_record(request, pk):
	if request.user.is_authenticated:
		# Look Up Records
		customer_record = Record.objects.get(id=pk)
		return render(request, 'record.html', {'customer_record':customer_record})
	else:
		messages.success(request, "Debe iniciar sesión para ver esa página...")
		return redirect('home')

def delete_record(request, pk):
	if request.user.is_authenticated:
		delete_it = Record.objects.get(id=pk)
		delete_it.delete()
		messages.success(request, "Registro eliminado con éxito...")
		return redirect('home')
	else:
		messages.success(request, "Debe haber iniciado sesión para hacer eso...")
		return redirect('home')

def add_record(request):
	form = AddRecordForm(request.POST or None)
	if request.user.is_authenticated:
		if request.method == "POST":
			if form.is_valid():
				add_record = form.save()
				messages.success(request, "Registro añadido...")
				return redirect('home')
		return render(request, 'add_record.html', {'form':form})
	else:
		messages.success(request, "Debe iniciar sesión para ver esa página.....")
		return redirect('home')

def update_record(request, pk):
	if request.user.is_authenticated:
		current_record = Record.objects.get(id=pk)
		form = AddRecordForm(request.POST or None, instance=current_record)
		if form.is_valid():
			form.save()
			messages.success(request, "¡El registro ha sido actualizado!")
			return redirect('home')
		return render(request, 'update_record.html', {'form':form})
	else:
		messages.success(request, "Se actualizo...")
		return redirect('home')

@login_required
def generar_reporte_excel(request):
    queryset = Record.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
	# Agregar encabezados de columna y datos de los registros
    encabezados = [
        'Codigo Institucion Externa',
        'Codigo Senecyt',
        'Codigo Toma Fisica',
        'Codigo Anterior',
        'Bien',
        'Clase Bien',
        'Serie',
        'Modelo',
        'Color',
        'Material',
        'Marca',
        'Estado',
        'Usuario Final',
        'Nro. Cedula',
        'Ubicacion',
        'Edificio',
        'Custodio Administrativo',
        'Delegado 1',
        'Delegado 2',
        'Delegado 3',
        'Tipo Novedad',
        'Observaciones',
        'Nro. Acta Entrega Recepcion',
        'Nro. Acta Donacion Factura',
        'Valor Monetario',
        'Fecha Entrega Recepcion',
        'Fecha Creación',
    ]

    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = openpyxl.utils.get_column_letter(col_num)
        celda = sheet[f'{col_letra}1']
        celda.value = encabezado
        celda.font = openpyxl.styles.Font(bold=True)

    # Agregar datos de los registros
    for row_num, registro in enumerate(queryset, 2):
        sheet.cell(row=row_num, column=1, value=registro.codigo_institucion_externa)
        sheet.cell(row=row_num, column=2, value=registro.codigo_senecyt)
        sheet.cell(row=row_num, column=3, value=registro.codigo_toma_fisica)
        sheet.cell(row=row_num, column=4, value=registro.codigo_anterior)
        sheet.cell(row=row_num, column=5, value=registro.bien)
        sheet.cell(row=row_num, column=6, value=registro.clase_bien)
        sheet.cell(row=row_num, column=7, value=registro.serie)
        sheet.cell(row=row_num, column=8, value=registro.modelo)
        sheet.cell(row=row_num, column=9, value=registro.color)
        sheet.cell(row=row_num, column=10, value=registro.material)
        sheet.cell(row=row_num, column=11, value=registro.marca)
        sheet.cell(row=row_num, column=12, value=registro.estado)
        sheet.cell(row=row_num, column=13, value=registro.usuario_final)
        sheet.cell(row=row_num, column=14, value=registro.nro_cedula)
        sheet.cell(row=row_num, column=15, value=registro.ubicacion)
        sheet.cell(row=row_num, column=16, value=registro.edificio)
        sheet.cell(row=row_num, column=17, value=registro.custodio_administrativo)
        sheet.cell(row=row_num, column=18, value=registro.delegado_1)
        sheet.cell(row=row_num, column=19, value=registro.delegado_2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
    
    workbook.save(response)
    return response
